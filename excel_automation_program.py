#another way to open the excel file
import openpyxl as xl
#  to add a chart
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
        wb = xl.load_workbook(filename)
        # to access the excel file
        sheet = wb['Sheet1']
        #to select the row and column
        # cell = sheet['a1']
        #another method to open a particular row and column
        # cell = sheet.cell(1, 1)
        # print(cell.value)
        #to check how many row is present in the file

        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            corrected_price = (cell.value) * 0.9
            corrected_price_cell = sheet.cell(row, 4)
            corrected_price_cell.value = corrected_price
        # to save to a file or overwrite it
        values = Reference(sheet, 
                    min_row=2, 
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

        chart = BarChart()
        chart.add_data(values)
        # to select the location of where we want the chart to be 
        sheet.add_chart(chart, 'e2')

        wb.save(filename)
    
# print(sheet.max_row)

